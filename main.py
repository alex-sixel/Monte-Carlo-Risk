#!/usr/bin/env python

import os

import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pert
from pert import PERT


# given lower, mean(probable), upper values and number of iterations, generate and return random values
def triangular_distribution(lower, more_likely, upper, iterations):
    return np.random.triangular(lower, more_likely, upper, iterations)


def pert_distribution(lower, more_likely, upper):
    return PERT(lower, more_likely, upper)


# plot a line
def plot_pert_line(pert_series, iterations, legenda='PERT'):
    # plt.figure()
    sns.kdeplot(pert.rvs(iterations), label=legenda, legend=False, color='orange')
    # sns.distplot(pert.rvs(iterations), label=legenda)


# generates pert histogram
def plot_pert(pert_series, iterations):
    # print("pert_series type is: ", type(pert_series))
    # print("pert.rvs(iterations) type is: ", type(pert.rvs(iterations)))
    # print("pert.rvs(5) is: ", pert.rvs(5),"\n\n")
    # plt.figure()
    # colour #1db4cf turquoise 0394ad (darker)
    return plt.hist(pert.rvs(iterations), bins=150, density=True, color='#0394ad')


# draw a vertical --- red line at given location
def draw_vertical_line(posX, colour='red', legenda=""):
    plt.axvline(x=posX, color=colour, linestyle='--', linewidth=1.7, label=legenda)


# write a bold text at given location from (0,0) to (1,1)
def write_bold_text(posX, posY, text, colour='black'):
    plt.text(posX, posY, text, color=colour, fontweight='bold', fontsize=10, transform=plt.gcf().transFigure)


# get probability of X
def get_p_of_x(given_list, probability):
    probability = probability / 100

    # - 1, because 0 is also a position
    position = (probability * len(given_list)) - 1

    result = given_list[int(position)].astype(int)
    given_list = given_list.astype(int)

    return result


def save_as_png(address, name):
    # support .png .pdf .svg(ignores dpi) files
    # plt.savefig("docs/monte-carlo.pdf", dpi=150)
    address = address + "/" + name + ".png"
    # , bbox_inches='tight'
    # print(address)
    plt.savefig(address, dpi=150)


# save chart into worksheet
def save_into_worksheet(planilha, img_name, localizacao='A1'):
    # add file extension and -r(for renamed)
    # img_renamed = img_name + "-r.png"
    # inserir caminho
    img_name = img_name + ".png"

    # open original image, resize and save it as renamed file
    # img = Image.open(img_name)
    # img = img.resize((1024,768),Image.NEAREST)
    # img.save(img_renamed)

    worksheet = planilha["Graphical Analysis"]

    # Open renamed file, attach it to worksheet
    # img = openpyxl.drawing.image.Image(img_renamed)

    img = openpyxl.drawing.image.Image(img_name)
    img.anchor = localizacao
    worksheet.add_image(img)


# read Excel file and gets distribution and value intervals
def read_document(file):
    wb = openpyxl.load_workbook(file, keep_vba=True)
    sheet = wb["Risk Matrix"]
    last_row = len(list(sheet.rows))
    estimative_list = []

    for i in range(9, last_row):
        row = i
        id = sheet.cell(row=i, column=1)
        risk = sheet.cell(row=i, column=2)
        distribution = sheet.cell(row=i, column=10)
        lower_value = sheet.cell(row=i, column=11)
        likely_value = sheet.cell(row=i, column=12)
        upper_value = sheet.cell(row=i, column=13)
        if risk.value is None:
            break
        estimative_list.append(
            [row, id.value, risk.value, distribution.value, lower_value.value, likely_value.value,
             upper_value.value])

    return estimative_list


# open sheet and insert risks at columns P30, P60 and P80
def write_risk_to_excel(planilha, row, result30, result60, result80):
    worksheet = planilha["Risk Matrix"]

    worksheet.cell(row=row, column=15).value = result30
    worksheet.cell(row=row, column=16).value = result60
    worksheet.cell(row=row, column=17).value = result80


if __name__ == "__main__":
    # 1k, 5k or 10k
    # For smaller values, recommended value is 5K
    iterations = 5000

    directory = os.getcwd()
    path = directory + '/Monte-Carlo.xlsm'

    # points to a new file
    # planilha = openpyxl.load_workbook(path, keep_vba=True)
    planilha = openpyxl.load_workbook(path)

    estimative_list = read_document(path)
    # print("Full list: ", estimative_list)

    linha = 1
    for i in estimative_list:
        if i[3] == 'PERT':
            row = i[0]
            id = i[1]
            name = i[2]
            lower_value = i[4]
            more_likely = i[5]
            upper_value = i[6]

            # clean previous plot (avoid plot problems)
            plt.clf()

            # (13,7) = 1300x700
            plt.figure(figsize=(13, 7))  # set window size (x,y). Default is (6.4,4.8) #remove if want a 640x480
            plt.grid(alpha=0.75)

            # set the background grid to grey
            # plt.rcParams['axes.facecolor'] =  '#b3b3b3'
            plt.xlabel("Value", fontweight='bold')
            plt.ylabel("Density", fontweight='bold')

            pert = pert_distribution(lower_value, more_likely, upper_value)
            # generate random sequence between values stablished
            pert_list = pert.rvs(iterations)
            pert_list.sort()

            # print("lista ", lista.astype(int))
            result30 = get_p_of_x(pert_list, 30)
            result60 = get_p_of_x(pert_list, 60)
            result80 = get_p_of_x(pert_list, 80)

            write_risk_to_excel(planilha, row, result30, result60, result80)

            # set vertical lines + legend info
            ptext30 = "P(" + str(30) + "): " + str(result30)
            draw_vertical_line(result30, 'red', ptext30)
            ptext60 = "P(" + str(60) + "): " + str(result60)
            draw_vertical_line(result60, 'orange', ptext60)
            ptext80 = "P(" + str(80) + "): " + str(result80)
            draw_vertical_line(result80, 'blue', ptext80)

            pert_hist = plot_pert(pert, iterations)
            plot_pert_line(pert, iterations, legenda="PERT")
            # bbox_to_anchor=(1,0.5) #set legend out the figure. uncomment to keep it inside the figure

            # Customize chart's title
            if len(name) > 80:
                title = "ID " + str(id) + " - " + str(name[:80]) + "...\n\n"
            else:
                title = "ID " + str(id) + " - " + name + "\n\n"

            plt.title(title)
            plt.legend()

            # Optional: Save chart as png file in same directory as excel file
            img_name = "ID-" + str(id) + "-monte-carlo-pert"
            save_as_png(directory, img_name)

            # location = A1, A45, A90 ... and so on
            location = openpyxl.utils.cell.get_column_letter(1) + str(linha)
            save_into_worksheet(planilha, img_name, location)
            linha = linha + 55

        if i[3] == 'TRIANGULAR':
            row = i[0]
            id = i[1]
            name = i[2]
            lower_value = i[4]
            more_likely = i[5]
            upper_value = i[6]

            triangular = triangular_distribution(lower_value, more_likely, upper_value, iterations)
            triangular.sort()

            # get_p_of_x = get probability of X
            result30 = get_p_of_x(triangular, 30)
            result60 = get_p_of_x(triangular, 60)
            result80 = get_p_of_x(triangular, 80)

            write_risk_to_excel(planilha, row, result30, result60, result80)
            # planilha.save(path)

            # clean previous plot (avoid plot problems)
            plt.clf()
            # set window info
            plt.figure(figsize=(13, 7))  # set window size (x,y). Default is (6.4,4.8)
            # info about bins: https://docs.astropy.org/en/stable/visualization/histogram.html
            h = plt.hist(triangular, bins=200, density=True, label="Triangular")
            plt.grid(alpha=0.75)
            plt.xlabel("Value", fontweight='bold')
            plt.ylabel("Density", fontweight='bold')

            # Customize chart's title
            if len(name) > 80:
                title = "ID " + str(id) + " - " + str(name[:80]) + "...\n\n"
            else:
                title = "ID " + str(id) + " - " + name + "\n\n"
            # print("Title is: ", title)
            plt.title(title)

            # Set legend and draw vertical line in the chart
            ptext30 = "P(" + str(30) + "): " + str(result30)
            draw_vertical_line(result30, 'red', ptext30)

            ptext60 = "P(" + str(60) + "): " + str(result60)
            draw_vertical_line(result60, 'orange', ptext60)

            ptext80 = "P(" + str(80) + "): " + str(result80)
            draw_vertical_line(result80, 'blue', ptext80)

            # bbox_to_anchor=(1,0.5) set legend out the figure. uncomment to keep it inside the figure
            plt.legend()
            # plt.legend(bbox_to_anchor=(1,0.85))

            # Optional: Save chart as png file in same directory as excel file
            img_name = "ID-" + str(id) + "-monte-carlo-triangular"
            # filenames.append(filename)
            save_as_png(directory, img_name)

            location = openpyxl.utils.cell.get_column_letter(1) + str(linha)
            # paste image in the sheet and "jump" 55 rows to paste the next image
            save_into_worksheet(planilha, img_name, location)
            linha = linha + 55

    # no need to save at the file we're working as we only want a copy to have data + charts
    # planilha.save(path)
    # save a copy without macros
    planilha.save(directory + '/Monte-Carlo_result.xlsx')
    planilha.close()
